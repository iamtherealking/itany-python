
from urllib import request,parse
from bs4 import BeautifulSoup
from openpyxl import Workbook
import pymysql
import requests
import uuid



# 图示标签对应的表头
book_tag_titles = ["编号", "标签名", "父标签"]
# 图示对应的表头
book_titles = ["编号", "书名", "封面", "作者", "出版社", "出版日期", "价格", "评分"]

# 图书标签url
book_tag_url = "https://www.douban.com/tag/"

# 图书列表的url
book_url = "https://www.douban.com/tag/{}/book"


# 获取ip proxy
def __get_proxy__():
    return requests.get("http://127.0.0.1:5010/get").content


# 获取url对应的html字符串
def __get_html_bs__(url):
    try:
        proxy = request.ProxyHandler({"sock5": __get_proxy__()})
        opener = request.build_opener(proxy)
        request.install_opener(opener)
        html_text = request.urlopen(url).read().decode("utf-8")
        bs = BeautifulSoup(html_text, "html.parser")
        return bs
    except Exception as e:
        print(e)
        return None


# 爬取图书标签
def book_tag_spider():
    bs = __get_html_bs__(book_tag_url)
    result = []
    if bs is None:
        return result
    tags_content = bs.select("#content .article")[0]
    tags_h2 = tags_content.select("h2")
    for one_h2 in tags_h2:
        tag_name = one_h2.get_text().strip()
        id = str(uuid.uuid1())
        result.append({
            "tag_name": tag_name,
            "id": id,
            "parent_id": None
        })

        lis = one_h2.next_sibling.next_sibling.select("li")
        for one_li in lis:
            tag_name = one_li.a.get_text().strip()
            li_id = str(uuid.uuid1())
            result.append({
                "tag_name": tag_name,
                "id": li_id,
                "parent_id": id
            })

    return result


# 根据标签爬取图书列表
def book_spider(tag_list):
    book_list = []
    index = 0
    for tag in tag_list:

        index += 1
        print(tag['tag_name'],index,len(tag_list))
        real_url = book_url.format(parse.quote(tag['tag_name']))
        try:
            bs = __get_html_bs__(real_url)
        except:
            continue
        # html = open("douban.html",encoding="utf-8").read()
        # bs = BeautifulSoup(html,"html.parser")
        if bs is None:
            continue
        dls = bs.select("#content dl")
        for dl in dls:
            try:
                img_src = dl.dt.img.get("src")
                book_name = dl.dd.a.get_text().strip()
                book_id = dl.dd.a.get("href")
                book_id = book_id[book_id.index("subject/") + len("subject/"):book_id.rindex("/")]
                desc = dl.select(".desc")[0].get_text().strip()
                desc_arr = desc.split("/")
                rate = dl.select(".rating_nums")[0].get_text().strip()
                real_auth = ""
                for m in desc_arr[0:len(desc_arr)-3]:
                    real_auth = real_auth +  " " + m
                book_list.append({
                    "id": book_id,
                    "name": book_name,
                    "book_url": img_src,
                    "author": real_auth,
                    "publisher": desc_arr[-3].strip(),
                    "publishDate": desc_arr[-2].strip(),
                    "price": desc_arr[-1].strip(),
                    "rate": rate
                })
            except Exception as e:
                print(e)
                continue
    return book_list


# 导出标签到数据库
def export_tag_to_mysql(data_list):
    sql = '''
        insert into 
            t_book_tag
            (id,tagName,parentId)
        VALUES 
            ('{}','{}','{}')
    '''
    # 创建 连接信息的 字典
    config = {
        "host": "localhost",
        "port": 3306,  # 默认是3306，可以不写
        "user": "root",
        "passwd": "",
        "db": "python",
        "charset": "utf8"
    }
    db = pymysql.connect(**config)
    cursor = db.cursor()
    try:
        for d in data_list:
            cursor.execute(sql.format(d['id'],d['tag_name'],d['parent_id']))
        db.commit()
    except Exception as e:
        print(e)
        db.rollback()
    db.close()


# 导出书籍信息到mysql
def export_book_to_mysql(book_list):
    stored_ids = []
    sql = '''
            insert into 
                t_book
                (id, name, book_url, author, publisher, publishDate, price, rate)
            VALUES 
                ('{}','{}','{}','{}','{}','{}','{}','{}')
        '''
    # 创建 连接信息的 字典
    config = {
        "host": "localhost",
        "port": 3306,  # 默认是3306，可以不写
        "user": "root",
        "passwd": "",
        "db": "python",
        "charset": "utf8"
    }
    db = pymysql.connect(**config)
    cursor = db.cursor()
    try:
        for d in book_list:
            if d['id'] in stored_ids:
                continue
            cursor.execute(sql.format(
                d['id'], d['name'], d['book_url'], d['author'],
                d['publisher'], d['publishDate'], d['price'], d['rate']
            ))
            stored_ids.append(d['id'])
        db.commit()
    except Exception as e:
        print(e)
        db.rollback()
    db.close()


# 导出到excel
def export_to_excel(tag_list, book_list):
    wb = Workbook()

    bts = wb.create_sheet("图书标签数据", 0)
    bts.append(book_tag_titles)
    for tag in tag_list:
        bts.append([tag['id'], tag['tag_name'], tag['parent_id']])

    booksheet = wb.create_sheet("图书数据", 1)
    booksheet.append(book_titles)
    for b in book_list:
        booksheet.append(
            [
                b['id'], b['name'], b['book_url'],b['author'],
                b['publisher'], b['publishDate'], b['price'],b['rate']

            ]
        )

    wb.save("douban_export.xlsx")


tag_list = book_tag_spider()
export_tag_to_mysql(tag_list)

book_list = book_spider(tag_list)
export_book_to_mysql(book_list)


export_to_excel(tag_list, book_list)
